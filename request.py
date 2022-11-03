import requests
import json
import openpyxl as op
from openpyxl import load_workbook

# print(response.encoding)
# response.encoding="utf-8"    #更改为utf-8编码
# print(response.status_code)  # 打印状态码
# print(response.url)          # 打印请求url
# print(response.headers)      # 打印头信息
# print(response.cookies)      # 打印cookie信息
# #print(response.text)  #以字符串形式打印网页源码
# #print(response.content) #以字节流形式打印

# https://leetcode.cn/graphql/
# url = 'https://leetcode.cn/graphql'

#===============================================================
body = {
    "operationName":"leetbookAllSubjects",
    "variables":{},
    "query":"query leetbookAllSubjects {\n  leetbookAllSubjects {\n    name\n    slug\n    __typename\n  }\n}\n"
}
headers = {
    "content-type":"application/json",
    "referer":"https://leetcode.cn/leetbook"
}

response = requests.post('https://leetcode.cn/graphql',json = body,headers = headers)
# response.encoding = "utf-8"
print(response.status_code)
# print(response.reason)
# # print(response.content.decode("unicode_escape"))
# print(response.json())
# # print(response.headers)
# # print(response.content)
# # print(response.text)
r = response.json()
data = r['data']['leetbookAllSubjects']
# print(data)
# print(len(data))
def op_toExcel1(data,fileName):
    wb = op.Workbook()
    ws = wb['Sheet']
    ws.append(['name','slug','__typename'])
    for i in range(len(data)):
        d = data[i]['name'],data[i]['slug'],data[i]['__typename']
        ws.append(d)
    wb.save(fileName)
fileName = "leetcode.xlsx"
op_toExcel1(data,fileName)

#=====================================

#=========================取出表格数据生成字典=============
workbook = load_workbook(filename='leetcode.xlsx')
print(workbook.sheetnames)
sheet = workbook['Sheet']
sheetlist = {}
for i in range(2,8):
    cellkey = sheet['A'+ str(i)]
    cellvalue = sheet['B' + str(i)]
    sheetlist.update({cellkey.value:cellvalue.value})
    # print(cell.value)
print(sheetlist)
#=======================
def op_toExcel(data,fileName,ws):
    wb =  load_workbook(filename='leetcode2.xlsx')
    # wb = op.Workbook()
    # print(sheetname,'is ok')
    # ws = wb[sheetname]
    # wb.save(fileName)
    ws.append(['name','slug','nameTranslated','__typename'])
    for i in range(len(data)):
        d = data[i]['name'],data[i]['slug'],data[i]['nameTranslated'],data[i]['__typename']
        ws.append(d)
    wb.save(fileName)
wb = op.Workbook()
wb.save('leetcode2.xlsx')
headers = {
    "content-type":"application/json",
    "referer":"https://leetcode.cn/leetbook"
}
for name in sheetlist:
    ws = wb.create_sheet(name)
    wb.save('leetcode2.xlsx')
    body1 = {
        "operationName":"leetbookTopTags",
        "variables":{"size":100,"subjectSlug":sheetlist[name]},
        "query":"query leetbookTopTags($size: Int!, $subjectSlug: String) {\n  leetbookTopTags(size: $size, subjectSlug: $subjectSlug) {\n    name\n    slug\n    nameTranslated\n    __typename\n  }\n}\n"
    }

    response1 = requests.post('https://leetcode.cn/graphql',json = body1,headers = headers)
    # response1.encoding = "utf-8"
    print(response1.status_code)
    # print(response1.reason)
    # print(response1.json())
    r1 = response1.json()
    data2 = r1['data']['leetbookTopTags']
    print(data2)
    fileName = "leetcode2.xlsx"
    op_toExcel(data2,fileName,ws)
    # print(data2)
    # print(len(data2))
wb.save('leetcode2.xlsx')


